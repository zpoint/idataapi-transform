import re
import logging
from openpyxl import Workbook
from .BaseWriter import BaseWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

_warning = False


class XLSXWriter(BaseWriter):
    def __init__(self, config):
        global _warning
        super().__init__()
        self.config = config
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = config.title
        self.col_dict = dict()
        self.row = 2
        self.total_miss_count = 0
        self.success_count = 0
        if not _warning:
            logging.warning("XLSXWriter will actually write to file when __exit__ of XLSXWriter called")
            _warning = True

    def write(self, responses):
        miss_count = 0
        for each_response in responses:
            if self.config.expand:
                each_response = self.expand_dict(each_response, max_expand=self.config.expand)
            if self.config.filter:
                each_response = self.config.filter(each_response)
                if not each_response:
                    miss_count += 1
                    continue

            for key, value in each_response.items():
                if key not in self.col_dict:
                    self.col_dict[key] = len(self.col_dict) + 1
                    self.ws1.cell(row=1, column=self.col_dict[key], value=key)
                value = str(value) if value is not None else ""
                try:
                    self.ws1.cell(row=self.row, column=self.col_dict[key], value=value)
                except Exception:
                    new_value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)
                    logging.warning("row num: %d, key: %s, value: %s contains illegal characters, "
                                    "replaced illegal characters to: %s" % (self.row, key, value, new_value))
                    self.ws1.cell(row=self.row, column=self.col_dict[key], value=new_value)

            self.row += 1
            self.success_count += 1
        logging.info("%s write %d item, filtered %d item" % (self.config.filename, len(responses), miss_count))

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(filename=self.config.filename)
        self.wb.close()
        logging.info("%s write done, total filtered %d item, total write %d item" %
                     (self.config.filename, self.total_miss_count, self.success_count))

    def __enter__(self):
        return self
