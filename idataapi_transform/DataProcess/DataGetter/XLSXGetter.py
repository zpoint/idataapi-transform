import logging
from openpyxl import load_workbook
from .BaseGetter import BaseGetter


class XLSXGetter(BaseGetter):
    def __init__(self, config):
        super().__init__()
        self.config = config
        self.wb = load_workbook(filename=self.config.filename, read_only=True)
        if not self.wb.worksheets:
            raise ValueError("Empty file: %s" % (self.config.filename, ))
        self.sheet = self.wb.worksheets[self.config.sheet_index]
        self.row_iter = self.sheet.rows
        self.headers = self.generate_headers()

        self.max_row = self.sheet.max_row
        if self.config.max_limit and self.config.max_limit > self.max_row:
            self.max_row = self.config.max_limit + 1  # add first headers

        self.row_num = 0
        self.responses = list()
        self.curr_size = 0
        self.done = False
        self.miss_count = 0
        self.total_count = 0

    def init_val(self):
        self.row_num = 0
        self.responses = list()
        self.curr_size = 0
        self.done = False
        self.miss_count = 0
        self.total_count = 0

        self.row_iter = self.sheet.rows

    def __aiter__(self):
        return self

    async def __anext__(self):
        if self.done:
            logging.info("get source done: %s, total get %d items, total filtered: %d items" %
                         (self.config.filename, self.total_count, self.miss_count))
            self.init_val()
            raise StopAsyncIteration

        while self.row_num < self.max_row:
            if self.row_num == 0:
                self.row_num += 1
                continue

            self.row_num += 1
            self.total_count += 1
            row = self.get_next_row()
            if self.config.filter:
                row = self.config.filter(row)
                if not row:
                    self.miss_count += 1
                    continue
            self.responses.append(row)
            if len(self.responses) > self.config.per_limit:
                self.curr_size += len(self.responses)
                return self.clear_and_return()

        if self.responses:
            self.done = True
            return self.clear_and_return()

        logging.info("get source done: %s, total get %d items, total filtered: %d items" %
                     (self.config.filename, self.total_count, self.miss_count))
        self.init_val()
        raise StopAsyncIteration

    def generate_headers(self):
        keys = list()
        try:
            row = next(self.row_iter)
            for each in row:
                keys.append(each.value)
        except StopIteration:
            pass
        return keys

    def get_next_row(self):
        ret_item = dict()
        r = next(self.row_iter)
        for key, cell in zip(self.headers, r):
            ret_item[key] = cell.value
        return ret_item

    def __iter__(self):
        for row_num in range(self.max_row):
            if row_num == 0:
                continue

            row_num += 1
            self.total_count += 1
            row = self.get_next_row()
            if self.config.filter:
                row = self.config.filter(row)
                if not row:
                    self.miss_count += 1
                    continue
            self.responses.append(row)
            if len(self.responses) > self.config.per_limit:
                self.curr_size += len(self.responses)
                yield self.clear_and_return()

        if self.responses:
            yield self.responses

            logging.info("get source done: %s, total get %d items, total filtered: %d items" %
                         (self.config.filename, self.total_count, self.miss_count))
            self.init_val()

    def __del__(self):
        self.wb.close()

    def clear_and_return(self):
        resp = self.responses
        self.responses = list()
        return resp
