import re
import os
import logging
from openpyxl import Workbook, load_workbook
from .BaseWriter import BaseWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from ..Config.DefaultValue import DefaultVal

_warning = False


class XLSXWriter(BaseWriter):
    def __init__(self, config):
        global _warning
        super().__init__()
        self.config = config
        self.col_dict = dict()
        self.row = 2
        # headers
        self.header_generated = False
        self.file_already_exists = os.path.exists(self.config.filename)
        if "a" in self.config.mode and self.file_already_exists:
            self.wb = load_workbook(filename=self.config.filename, read_only=False)
            self.generate_header(from_file=True)
        else:
            self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = config.title
        self.total_miss_count = 0
        self.success_count = 0
        if not _warning:
            logging.warning("XLSXWriter will actually write to file when __exit__ of XLSXWriter called")
            _warning = True

    def write(self, responses):
        if not self.header_generated and self.config.headers:
            self.generate_header()

        miss_count = 0
        for each_response in responses:
            if self.config.expand:
                each_response = self.expand_dict(each_response, max_expand=self.config.expand)
            if self.config.filter:
                each_response = self.config.filter(each_response)
                if not each_response:
                    miss_count += 1
                    continue

            for key, value in each_response.items():
                if key not in self.col_dict:
                    self.col_dict[key] = len(self.col_dict) + 1
                    self.ws1.cell(row=1, column=self.col_dict[key], value=key)
                value = str(value) if value is not None else ""
                try:
                    self.ws1.cell(row=self.row, column=self.col_dict[key], value=value)
                except Exception:
                    new_value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)
                    logging.warning("row num: %d, key: %s, value: %s contains illegal characters, "
                                    "replaced illegal characters to: %s" % (self.row, key, value, new_value))
                    self.ws1.cell(row=self.row, column=self.col_dict[key], value=new_value)

            self.row += 1
            self.success_count += 1
        logging.info("%s write %d item, filtered %d item" % (self.config.filename, len(responses), miss_count))

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(filename=self.config.filename)
        self.wb.close()
        logging.info("%s write done, total filtered %d item, total write %d item" %
                     (self.config.filename, self.total_miss_count, self.success_count))

    def __enter__(self):
        return self

    def generate_header(self, from_file=False):
        if from_file:
            if not self.wb.worksheets:
                # empty file
                return
            sheet = self.wb.worksheets[self.config.sheet_index]
            row_iter = sheet.rows
            try:
                row = next(row_iter)
                for each in row:
                    self.col_dict[each.value] = len(self.col_dict) + 1
            except StopIteration:
                # empty file
                return
            if len(self.col_dict) == 1 and list(self.col_dict.keys())[0] is None:
                # empty file
                self.col_dict.clear()
                return
            max_row = sheet.max_row
            self.row = max_row + 1
        else:
            for key in self.config.headers:
                self.col_dict[key] = len(self.col_dict) + 1
                self.ws1.cell(row=1, column=self.col_dict[key], value=key)

        self.header_generated = True
