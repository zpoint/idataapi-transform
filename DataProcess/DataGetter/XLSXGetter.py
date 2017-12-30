import logging
from openpyxl import load_workbook
from .BaseGetter import BaseGetter


class XLSXGetter(BaseGetter):
    def __init__(self, config):
        super().__init__()
        self.config = config
        self.wb = load_workbook(filename=self.config.filename, read_only=True)
        if not self.wb.worksheets:
            raise ValueError("Empty file: %s" % (self.config.filename, ))
        self.sheet = self.wb.worksheets[0]
        self.headers = self.generate_headers()

        self.max_row = self.sheet.max_row
        if self.config.max_limit and self.config.max_limit > self.max_row:
            self.max_row = self.config.max_limit + 1  # add first headers

        self.row_num = 0
        self.responses = list()
        self.curr_size = 0
        self.need_clear = False
        self.done = False

    def init_val(self):
        self.row_num = 0
        self.responses = list()
        self.curr_size = 0
        self.need_clear = False
        self.done = False

    def __aiter__(self):
        return self

    def __anext__(self):
        if self.need_clear:
            self.responses.clear()
            self.need_clear = False

        if self.done:
            logging.info("get source done: %s" % (self.config.filename,))
            self.init_val()
            raise StopAsyncIteration

        while self.row_num < self.max_row:
            if self.row_num == 0:
                continue

            self.row_num += 1
            row = self.get_row(self.row_num)
            self.responses.append(row)
            if len(self.responses) > self.config.per_limit:
                self.curr_size += len(self.responses)
                self.need_clear = True
                return self.responses

        if self.responses:
            self.need_clear = self.done = True
            return self.responses

        logging.info("get source done: %s" % (self.config.filename,))
        self.init_val()
        raise StopAsyncIteration

    def generate_headers(self):
        keys = list()
        for index in range(self.sheet.max_column):
            cell_index = index + 1
            keys.append(self.sheet._get_cell(1, cell_index).value)
        return keys

    def get_row(self, row_num):
        item = dict()
        for index in range(len(self.headers)):
            item[self.headers[index]] = self.sheet._get_cell(row_num, index+1).value
        return item

    def __iter__(self):
        for row_num in range(self.max_row):
            if row_num == 0:
                continue

            row_num += 1
            row = self.get_row(row_num)
            self.responses.append(row)
            if len(self.responses) > self.config.per_limit:
                self.curr_size += len(self.responses)
                yield self.responses
                self.responses.clear()

        if self.responses:
            yield self.responses
            self.init_val()
            logging.info("get source done: %s" % (self.config.filename,))

    def __del__(self):
        self.wb.close()
